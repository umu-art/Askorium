import logging
from io import BytesIO

import httpx

logger = logging.getLogger("document_renderer")


class DocumentRenderer:
    def __init__(self, timeout_s: int = 30):
        self._timeout = timeout_s

    async def render(self, url: str, hint: str) -> str:
        data = await self._download(url)
        hint = hint.lower()
        if hint == "pdf":
            return _render_pdf(data)
        elif hint in ("doc", "docx"):
            return _render_docx(data)
        elif hint in ("xls", "xlsx"):
            return _render_xlsx(data)
        raise ValueError(f"unsupported document type: {hint}")

    async def _download(self, url: str) -> bytes:
        async with httpx.AsyncClient(timeout=self._timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content


def _render_pdf(data: bytes) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(stream=data, filetype="pdf")
    parts = []
    for page in doc:
        blocks = page.get_text("dict")["blocks"]
        sizes = [
            span["size"]
            for b in blocks if b["type"] == 0
            for line in b["lines"]
            for span in line["spans"]
            if span["text"].strip()
        ]
        avg_size = sum(sizes) / len(sizes) if sizes else 12

        for b in blocks:
            if b["type"] != 0:
                continue
            for line in b["lines"]:
                text = " ".join(s["text"] for s in line["spans"]).strip()
                if not text:
                    continue
                max_size = max((s["size"] for s in line["spans"]), default=12)
                tag = "h2" if max_size >= avg_size * 1.3 else "p"
                parts.append(f"<{tag}>{text}</{tag}>")

    doc.close()
    return f"<html><body>{''.join(parts)}</body></html>"


def _render_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(data))
    parts = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = para.style.name.lower() if para.style else ""
        if "heading" in style:
            level = next((c for c in style if c.isdigit()), "2")
            parts.append(f"<h{level}>{text}</h{level}>")
        elif "list" in style:
            parts.append(f"<li>{text}</li>")
        else:
            parts.append(f"<p>{text}</p>")
    return f"<html><body>{''.join(parts)}</body></html>"


def _render_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"<h2>{sheet.title}</h2><table>")
        for row in sheet.iter_rows(values_only=True):
            cells = "".join(f"<td>{c if c is not None else ''}</td>" for c in row)
            parts.append(f"<tr>{cells}</tr>")
        parts.append("</table>")
    return f"<html><body>{''.join(parts)}</body></html>"
